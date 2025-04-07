import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])
import torch
from torch.utils.data import DataLoader
import torch.nn as nn
import torch.optim as optim
import os
import yaml
import ssl
import pandas as pd
import collections
from torchmetrics import PeakSignalNoiseRatio
from src.model.CycleGAN.discriminator_model import Discriminator
from src.model.CycleGAN.generator_model import Generator
import src.utils.utils as util_general
import src.utils.util_data as util_data
import src.model.CycleGAN.util_model as util_model
from openpyxl import Workbook
from openpyxl import load_workbook
from sys import platform

torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

# Configuration file

if platform == 'win32': # su windows
    args = {}
    args['cfg_file'] = "./configs/cyclegan_train.yaml"
    with open(args['cfg_file']) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)
else: # su alvis
    args = util_general.get_args()
    with open(args.cfg_file) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)


# Parameters
exp_name = cfg['exp_name']
model_name = cfg['model']['model_name']
learning_rate = cfg['trainer']['optimizer']['lr']
cv = cfg['data']['cv']
fold_list = list(range(cv))

# Device
device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print(device)

# Files and Directories
fold_public_dir = os.path.join(cfg['data']['fold_public_dir'])

model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

report_file_mse = os.path.join(report_dir, 'report_mse.xlsx')
report_file_psnr = os.path.join(report_dir, 'report_psnr.xlsx')
report_file_vif = os.path.join(report_dir, 'report_vif.xlsx')
report_file_ssim = os.path.join(report_dir, 'report_ssim.xlsx')

plot_training_dir = os.path.join(report_dir, "training")
util_general.create_dir(plot_training_dir)

outputs_dir = os.path.join(report_dir, "outputs")
util_general.create_dir(outputs_dir)

# CV
results_mse = collections.defaultdict(lambda: [])
metric_cols_mse = []
results_psnr = collections.defaultdict(lambda: [])
metric_cols_psnr = []
results_vif = collections.defaultdict(lambda: [])
metric_cols_vif = []
results_ssim = collections.defaultdict(lambda: [])
metric_cols_ssim = []
for fold in fold_list:
    # Dir
    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)
    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(outputs_fold_dir)

    # Results Frame
    metric_cols_mse.append("%s MSE" % str(fold))
    metric_cols_psnr.append("%s PSNR" % str(fold))
    metric_cols_vif.append("%s VIF" % str(fold))
    metric_cols_ssim.append("%s SSIM" % str(fold))

    # Data Loaders - Public Dataset
    fold_data = {step: pd.read_csv(os.path.join(fold_public_dir, str(fold), '%s.txt' % step), delimiter="\t", index_col='id') for step in ['train', 'val', 'test']}
    datasets = {step: util_data.ImgDataset_public_dataset(data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True) for step in ['train', 'val', 'test']}
    data_loaders = {'train': torch.utils.data.DataLoader(datasets['train'], batch_size=cfg['data']['batch_size'], shuffle=True, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
                    'val': torch.utils.data.DataLoader(datasets['val'], batch_size=cfg['data']['batch_size'], shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
                    'test': torch.utils.data.DataLoader(datasets['test'], batch_size=cfg['data']['batch_size'], shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker)}

    # Creo file excel con info generali
    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        # Definisco le colonne del file excel
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name'
        sheet['A2'] = exp_name
        sheet['B1'] = 'batch size'
        sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs'
        sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim'
        sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)

    # Inizializzo modelli
    disc_REC = Discriminator(in_channels=1).to(device)
    disc_LE = Discriminator(in_channels=1).to(device)
    gen_LE = Generator(img_channels=1, num_residuals=9).to(device)
    gen_REC = Generator(img_channels=1, num_residuals=9).to(device)
    opt_disc = optim.Adam(
        list(disc_REC.parameters()) + list(disc_LE.parameters()),
        lr=learning_rate,
        betas=(0.5, 0.999), # 0.5 per il momento, 0.999 per beta
    )

    opt_gen = optim.Adam(
        list(gen_LE.parameters()) + list(gen_REC.parameters()),
        lr=learning_rate,
        betas=(0.5, 0.999),
    )

    L1 = nn.L1Loss().to(device) # L1 --> cycle consistency loss e identity loss
    mse = nn.MSELoss().to(device) # mse --> adversarial loss
    early_stopping_criterion = nn.L1Loss().to(device)

    # Train model
    gen_REC, gen_LE, disc_REC, disc_LE, history = util_model.train_cycle_gan(gen_REC=gen_REC, gen_LE=gen_LE, disc_REC=disc_REC, disc_LE=disc_LE,
                                                                             data_loaders=data_loaders, early_stopping_criterion=early_stopping_criterion,
                                                                             opt_disc=opt_disc, opt_gen=opt_gen, L1=L1, mse=mse,
                                                                             model_fold_dir=model_fold_dir, cfg_trainer=cfg['trainer'], device=device)


    # Plot Training
    util_model.plot_training(history=history, model_name=model_name, plot_training_dir=plot_training_fold_dir)

    # Test model
    mse = nn.MSELoss()
    psnr = PeakSignalNoiseRatio().to(device)

    mse_test_results, psnr_test_results, vif_test_results, ssim_test_results = util_model.test_4metrics(gen_REC, loader_train=False, loader_val=False, loader_test=data_loaders['test'], mse_metric=mse, psnr_metric=psnr, device=device, outputs_dir=outputs_fold_dir, save_output=True)

    # Update report
    results_mse["%s MSE" % str(fold)].append(mse_test_results)
    results_psnr["%s PSNR" % str(fold)].append(psnr_test_results)
    results_vif["%s VIF" % str(fold)].append(vif_test_results)
    results_ssim["%s SSIM" % str(fold)].append(ssim_test_results)

    # Save Results MSE
    results_mse_frame = pd.DataFrame(results_mse)
    results_mse_frame.insert(loc=0, column='std MSE', value=results_mse_frame[metric_cols_mse].std(axis=1))
    results_mse_frame.insert(loc=0, column='mean MSE', value=results_mse_frame[metric_cols_mse].mean(axis=1))
    results_mse_frame.insert(loc=0, column='model', value=model_name)
    results_mse_frame.to_excel(report_file_mse, index=False)

    # Save Results PSNR
    results_psnr_frame = pd.DataFrame(results_psnr)
    results_psnr_frame.insert(loc=0, column='std PSNR', value=results_psnr_frame[metric_cols_psnr].std(axis=1))
    results_psnr_frame.insert(loc=0, column='mean PSNR', value=results_psnr_frame[metric_cols_psnr].mean(axis=1))
    results_psnr_frame.insert(loc=0, column='model', value=model_name)
    results_psnr_frame.to_excel(report_file_psnr, index=False)

    # Save Results VIF
    results_vif_frame = pd.DataFrame(results_vif)
    results_vif_frame.insert(loc=0, column='std VIF', value=results_vif_frame[metric_cols_vif].std(axis=1))
    results_vif_frame.insert(loc=0, column='mean VIF', value=results_vif_frame[metric_cols_vif].mean(axis=1))
    results_vif_frame.insert(loc=0, column='model', value=model_name)
    results_vif_frame.to_excel(report_file_vif, index=False)

    # Save Results SSIM
    results_ssim_frame = pd.DataFrame(results_ssim)
    results_ssim_frame.insert(loc=0, column='std SSIM', value=results_ssim_frame[metric_cols_ssim].std(axis=1))
    results_ssim_frame.insert(loc=0, column='mean SSIM', value=results_ssim_frame[metric_cols_ssim].mean(axis=1))
    results_ssim_frame.insert(loc=0, column='model', value=model_name)
    results_ssim_frame.to_excel(report_file_ssim, index=False)
