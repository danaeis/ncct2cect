import sys;

print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])
import os
import torch
import torch.nn as nn
import torch.optim as optim
from torch.optim import lr_scheduler
import pandas as pd
import collections
import yaml
import ssl
from sys import platform

import src.utils.utils as util_general
import src.utils.util_data as util_data
import src.model.Autoencoder.util_model as util_model

from openpyxl import Workbook
from openpyxl import load_workbook
from torchmetrics import PeakSignalNoiseRatio

torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

# Configuration file
if platform == 'win32': # su windows
    args = {}
    args['cfg_file'] = "./configs/autoencoder_tran_learn.yaml"
    with open(args['cfg_file']) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)
else: # su alvis
    args = util_general.get_args()
    with open(args.cfg_file) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)

# Seed everything
util_general.seed_all(cfg['seed'])

# Parameters
exp_name = cfg['exp_name']
model_name = cfg['model']['model_name']
cv = cfg['data']['cv']
fold_list = list(range(cv))

# Device
device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print(device)

# Files and Directories
fold_dir = os.path.join(cfg['data']['fold_dir'])  # folds (dentro processed)

model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

report_file_mse = os.path.join(report_dir, 'report_mse.xlsx')
report_file_psnr = os.path.join(report_dir, 'report_psnr.xlsx')
report_file_vif = os.path.join(report_dir, 'report_vif.xlsx')
report_file_ssim = os.path.join(report_dir, 'report_ssim.xlsx')

plot_training_dir = os.path.join(report_dir, "training")
util_general.create_dir(plot_training_dir)

outputs_dir = os.path.join(report_dir, "outputs")
util_general.create_dir(outputs_dir)

# CV
results_mse = collections.defaultdict(lambda: [])
metric_cols_mse = []
results_psnr = collections.defaultdict(lambda: [])
metric_cols_psnr = []
results_vif = collections.defaultdict(lambda: [])
metric_cols_vif = []
results_ssim = collections.defaultdict(lambda: [])
metric_cols_ssim = []

for fold in fold_list:
    # Dir
    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)
    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(outputs_fold_dir)

    # Results Frame
    metric_cols_mse.append("%s MSE" % str(fold))
    metric_cols_psnr.append("%s PSNR" % str(fold))
    metric_cols_vif.append("%s VIF" % str(fold))
    metric_cols_ssim.append("%s SSIM" % str(fold))

    # Data Loaders - FPUCBM Dataset
    fold_data = {step: pd.read_csv(os.path.join(fold_dir, str(fold), '%s.txt' % step), delimiter="\t", index_col='id') for step in ['train', 'val', 'test']}
    datasets = {step: util_data.ImgDataset(data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True) for step in ['train', 'val', 'test']}
    data_loaders = {'train': torch.utils.data.DataLoader(datasets['train'], batch_size=cfg['data']['batch_size'], shuffle=True, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
                    'val': torch.utils.data.DataLoader(datasets['val'], batch_size=cfg['data']['batch_size'], shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
                    'test': torch.utils.data.DataLoader(datasets['test'], batch_size=cfg['data']['batch_size'], shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker)}

    # Creo file excel
    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        # Definisco le colonne del file excel
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name'
        sheet['A2'] = exp_name
        sheet['B1'] = 'batch size'
        sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs'
        sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img dim'
        sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)

    # Model
    print("%s%s%s" % ("*" * 50, model_name, "*" * 50))

    # Carico modello addestrato su dataset pubblico
    pretrained_model = os.path.join(cfg['data']['model_dir'], cfg['model']['pretrained_model'])
    pretrained_model_dir = os.path.join(pretrained_model, str(fold))
    CHECKPOINT = os.path.join(pretrained_model_dir, "autoencoder_gong.pt")

    model = util_model.initialize_model(model_name=model_name, cfg_model=cfg['model'], device=device)
    checkpoint = torch.load(CHECKPOINT, map_location=device).module  # map_location = device
    model.load_state_dict(checkpoint.state_dict())

    if torch.cuda.device_count() > 1:
        print("Let's use", torch.cuda.device_count(), "GPUs!")
        model = nn.DataParallel(model)
    model = model.to(device)

    # Loss function
    criterion = nn.L1Loss().to(device)
    # Optimizer
    optimizer = optim.Adam(model.parameters(), lr=cfg['trainer']['optimizer']['lr'], weight_decay=cfg['trainer']['optimizer']['weight_decay'])
    # LR Scheduler
    scheduler = lr_scheduler.ReduceLROnPlateau(optimizer, mode=cfg['trainer']['scheduler']['mode'], patience=cfg['trainer']['scheduler']['patience'])
    # Train model
    model, history = util_model.train_model(model=model, criterion=criterion, optimizer=optimizer, scheduler=scheduler, model_name=model_name, data_loaders=data_loaders, model_dir=model_fold_dir, device=device, cfg_trainer=cfg['trainer'])

    # Plot Training
    util_model.plot_training(history=history, model_name=model_name, plot_training_dir=plot_training_fold_dir)

    # Test model
    mse = nn.MSELoss()
    psnr = PeakSignalNoiseRatio().to(device)

    mse_test_results, psnr_test_results, vif_test_results, ssim_test_results = util_model.test_4metrics(model, loader_train=False, loader_val=False, loader_test=data_loaders['test'], mse_metric=mse, psnr_metric=psnr,device=device, outputs_dir=outputs_fold_dir, save_output=True)

    # Update report
    results_mse["%s MSE" % str(fold)].append(mse_test_results)
    results_psnr["%s PSNR" % str(fold)].append(psnr_test_results)
    results_vif["%s VIF" % str(fold)].append(vif_test_results)
    results_ssim["%s SSIM" % str(fold)].append(ssim_test_results)

    # Save Results MSE
    results_mse_frame = pd.DataFrame(results_mse)
    results_mse_frame.insert(loc=0, column='std MSE', value=results_mse_frame[metric_cols_mse].std(axis=1))
    results_mse_frame.insert(loc=0, column='mean MSE', value=results_mse_frame[metric_cols_mse].mean(axis=1))
    results_mse_frame.insert(loc=0, column='model', value=model_name)
    results_mse_frame.to_excel(report_file_mse, index=False)

    # Save Results PSNR
    results_psnr_frame = pd.DataFrame(results_psnr)
    results_psnr_frame.insert(loc=0, column='std PSNR', value=results_psnr_frame[metric_cols_psnr].std(axis=1))
    results_psnr_frame.insert(loc=0, column='mean PSNR', value=results_psnr_frame[metric_cols_psnr].mean(axis=1))
    results_psnr_frame.insert(loc=0, column='model', value=model_name)
    results_psnr_frame.to_excel(report_file_psnr, index=False)

    # Save Results VIF
    results_vif_frame = pd.DataFrame(results_vif)
    results_vif_frame.insert(loc=0, column='std VIF', value=results_vif_frame[metric_cols_vif].std(axis=1))
    results_vif_frame.insert(loc=0, column='mean VIF', value=results_vif_frame[metric_cols_vif].mean(axis=1))
    results_vif_frame.insert(loc=0, column='model', value=model_name)
    results_vif_frame.to_excel(report_file_vif, index=False)

    # Save Results SSIM
    results_ssim_frame = pd.DataFrame(results_ssim)
    results_ssim_frame.insert(loc=0, column='std SSIM', value=results_ssim_frame[metric_cols_ssim].std(axis=1))
    results_ssim_frame.insert(loc=0, column='mean SSIM', value=results_ssim_frame[metric_cols_ssim].mean(axis=1))
    results_ssim_frame.insert(loc=0, column='model', value=model_name)
    results_ssim_frame.to_excel(report_file_ssim, index=False)
